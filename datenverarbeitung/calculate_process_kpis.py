import os
import pandas as pd
import csv
import openpyxl
import datetime
from datetime import datetime
from datetime import timedelta

# Make one function for each KPI and then a "main" function at the end that
# just calls all of them in order and updates the KPIs table

"""
Timestamp functions
"""


def add_timestamps(a, b):
    # Convert the duration strings to timedelta objects
    a_timedelta = timedelta(hours=int(a[:2]), minutes=int(a[3:5]), seconds=int(a[6:8]))
    b_timedelta = timedelta(hours=int(b[:2]), minutes=int(b[3:5]), seconds=int(b[6:8]))

    # Add the timedelta objects
    result_timedelta = a_timedelta + b_timedelta

    # Format the result as a string in hh:mm:ss format
    hours, remainder = divmod(result_timedelta.total_seconds(), 3600)
    minutes, seconds = divmod(remainder, 60)
    result_string = f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"

    return result_string


def subtract_timestamps(a, b):
    # Convert the duration strings to timedelta objects
    a_timedelta = timedelta(hours=int(a[:2]), minutes=int(a[3:5]), seconds=int(a[6:8]))
    b_timedelta = timedelta(hours=int(b[:2]), minutes=int(b[3:5]), seconds=int(b[6:8]))

    # Subtract the timedelta objects
    result_timedelta = a_timedelta - b_timedelta

    # Format the result as a string in hh:mm:ss format
    hours, remainder = divmod(result_timedelta.total_seconds(), 3600)
    minutes, seconds = divmod(remainder, 60)
    result_string = f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"

    return result_string


def divide_time(s, n):
    # Convert the duration string to a timedelta object
    s_timedelta = timedelta(hours=int(s[:2]), minutes=int(s[3:5]), seconds=int(s[6:8]))

    # Divide the timedelta by n, round it down to the nearest second
    result_timedelta = s_timedelta / n
    result_timedelta = timedelta(seconds=int(result_timedelta.total_seconds()))

    # Format the result as a string in hh:mm:ss format
    hours, remainder = divmod(result_timedelta.total_seconds(), 3600)
    minutes, seconds = divmod(remainder, 60)
    result_string = f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"

    return result_string


def average_time(time_list):
    # Ensure the input list is not empty
    if not time_list:
        return "00:00:00"

    # Initialize total_duration as the first timestamp in the list
    total_duration = time_list[0]

    # Iterate through the rest of the timestamps and add their durations
    for timestamp in time_list[1:]:
        total_duration = add_timestamps(total_duration, timestamp)

    # Divide the total duration by the number of timestamps
    average_duration = divide_time(total_duration, len(time_list))

    return average_duration


"""
KPI functions
"""


def calculate_average_cycle_time(variant, process, process_df, timestamp, period="day"):
    # Period to search for in the logs
    search_interval = 86400
    if period == "week":
        search_interval = 86400 * 7
    elif period == "month":
        search_interval = 86400 * 30
    elif period == "year":
        search_interval = 86400 * 365

    # Day of the last instance of the process
    today = process_df.tail(1)["date"].iloc[0]
    print(today)

    # Filter events that happened within the desired period
    filtered_process_df = process_df[
        (process_df["date"] == today) & (process_df["variant"] == variant) & (process_df["exit_code"] == 0)]

    # Use the durations to calculate the mean, divide by the Losgroesse to get cycle time per unit
    print(filtered_process_df["duration"])
    durations = filtered_process_df["duration"].tolist()
    print(durations)

    cycle_time = divide_time(average_time(durations), filtered_process_df["menge"].mean())

    current_directory = os.path.dirname(os.path.abspath(__file__))
    process_folder = os.path.join(current_directory, '..', 'Werk', 'Prozesse', process)
    file_path = os.path.join(process_folder, f'{process}_DS.csv')

    # Formatted variant
    translate = {"FK1": "KS_1", "FK2": "KS_2", "FK3": "KS_3", "FK4": "KS_4", "FK5": "KS_5", "FK6": "KS_6",
                 "FK7": "KS_7", "FK8": "KS_8", "FK": "KS (avg)", "FB1": "D25", "FB2": "D40", "FS1": "D25", "FS2": "D40"}

    variant = 'Cycle Time ' + f'{translate[variant]}'
    time_obj = datetime.strptime(cycle_time, "%H:%M:%S")
    cycle_time = time_obj.hour * 3600 + time_obj.minute * 60 + time_obj.second
    write_kpi(file_path, variant, cycle_time)

    return cycle_time


def calculate_process_time(process, variant, process_df, timestamp):
    # Day of the last instance of the process
    today = process_df.tail(1)["date"].iloc[0]

    # Filter events that happened within the desired period
    filtered_process_df = process_df[
        (process_df["date"] == today) & (process_df["variant"] == variant) & (process_df["exit_code"] == 0)]

    # Use the durations to calculate the mean, divide by the Losgroesse to get cycle time per unit
    durations = filtered_process_df["duration"].tolist()
    process_time = average_time(durations)

    # Calculate average cycle time
    current_directory = os.path.dirname(os.path.abspath(__file__))
    process_folder = os.path.join(current_directory, '..', 'Werk', 'Prozesse', process)
    file_path = os.path.join(process_folder, f'{process}_DS.csv')

    # Formatted variant
    translate = {"FK1": "KS_1", "FK2": "KS_2", "FK3": "KS_3", "FK4": "KS_4", "FK5": "KS_5", "FK6": "KS_6",
                 "FK7": "KS_7", "FK8": "KS_8", "FK": "KS (avg)", "FB1": "D25", "FB2": "D40", "FS1": "D25", "FS2": "D40"}

    variant = 'Prozesszeit ' + f'{translate[variant]}'
    time_obj = datetime.strptime(process_time, "%H:%M:%S")
    process_time = time_obj.hour * 3600 + time_obj.minute * 60 + time_obj.second
    write_kpi(file_path, variant, process_time)

    return process_time


def calculate_production_downtime(process, fehler_excel_path):
    ausfall_df = pd.read_excel(fehler_excel_path, sheet_name='LogData(Ausfallzeit)')
    try:
        today = ausfall_df.tail(1)["Datum"].iloc[0]
        filtered_df = ausfall_df[ausfall_df["Datum"] == today]
    except:
        filtered_df = ausfall_df

    downtime = "00:00:00"
    translate_machine = {"Drehen": "Drehmaschine", "Fraesen": "Fraesmaschine", "Saegen": "Bandsaege"}
    if process in ["Drehen", "Fraesen", "Saegen"]:
        filtered_filtered_df = filtered_df[filtered_df["Maschine"] == translate_machine[process]]
        number_of_rows = filtered_filtered_df.shape[0]
        for i in range(number_of_rows):
            downtime_duration = subtract_timestamps(str(filtered_filtered_df.iloc[i]["End"])
                                                    , str(filtered_filtered_df.iloc[i]["Start"]))
            print(downtime_duration)
            downtime = add_timestamps(downtime, downtime_duration)

    current_directory = os.path.dirname(os.path.abspath(__file__))
    process_folder = os.path.join(current_directory, '..', 'Werk', 'Prozesse', process)
    file_path = os.path.join(process_folder, f'{process}_DS.csv')
    variant = 'Production Downtime'
    write_kpi(file_path, variant, downtime)

    return downtime


def calculate_unscheduled_downtime(process, fehler_excel_path):
    df = pd.read_excel(fehler_excel_path, sheet_name='LogData(Ausfallzeit)')

    try:
        # Ensure the 'Start', 'End', 'Date', and 'Process' columns exist in the DataFrame
        if 'Start' in df.columns and 'End' in df.columns and 'Datum' in df.columns and 'Maschine' in df.columns:
            # Convert 'Date' column to datetime format (if not already)
            if not pd.api.types.is_datetime64_ns_dtype(df['Datum']):
                df['Datum'] = pd.to_datetime(df['Datum'])

            # Filter rows where 'Start' and 'End' are on the same date and match the desired process
            same_date_and_process_rows = df[(df['Start'].dt.date == df['End'].dt.date) & (df['Maschine'] == process)]

            # Calculate the sum of 'End' - 'Start' for these rows
            total_duration = same_date_and_process_rows['End'] - same_date_and_process_rows['Start']
            total_duration = total_duration.sum()

            return total_duration
        else:
            return None
    except:
        return 0


def calculate_nacharbeitquote(process, variant, process_df, fehler_excel_path):  # Prozentzahl
    today = process_df.tail(1)["date"].iloc[0]

    # Filter the DataFrame for the desired variant and today's date
    filtered_df = process_df[
        (process_df['variant'] == variant) & (process_df['date'] == today) & (process_df['exit_code'] == 2)]

    # Count the number of rows in the filtered DataFrame
    variants_today = filtered_df.shape[0]
    batch = process_df["menge"].tail(1).iloc[0]
    total_parts = variants_today * batch

    fehler_excel = pd.read_excel(fehler_excel_path, sheet_name='LogData(Fertigung)')

    try:
        for row in fehler_excel.iterrows():
            for cell in row:
                if cell.value == process:
                    process_cell = cell
                    nacharbeitsteil_cell = fehler_excel.cell(row=process_cell.row + 2, column=process_cell.column + 1)
                    na_teile = nacharbeitsteil_cell.value

                    nacharbeitsquote = na_teile * 100 / total_parts

                    current_directory = os.path.dirname(os.path.abspath(__file__))
                    process_folder = os.path.join(current_directory, '..', 'Werk', 'Prozesse', process)
                    file_path = os.path.join(process_folder, f'{process}_DS.csv')
                    variant = 'Fehlersproduktionsquote'
                    write_kpi(file_path, variant, nacharbeitsquote)

                    return nacharbeitsquote
    except:
        return 0


def calculate_ausschussquote(process, variant, process_df, fehler_excel_path):  # Prozentzahl
    today = process_df.tail(1)["date"].iloc[0]

    # Filter the DataFrame for the desired variant and today's date
    filtered_df = process_df[(process_df['variant'] == variant) & (process_df['date'] == today)]

    # Count the number of rows in the filtered DataFrame
    variants_today = filtered_df.shape[0]
    batch = process_df["menge"].tail(1).iloc[0]
    total_parts = variants_today * batch

    fehler_excel = pd.read_excel(fehler_excel_path, sheet_name='LogData(Fertigung)')

    try:
        for row in fehler_excel.iterrows():
            for cell in row:
                if cell.value == process:
                    process_cell = cell
                    ausschissteil_cell = fehler_excel.cell(row=process_cell.row + 2, column=process_cell.column)
                    as_teile = ausschissteil_cell.value
                    ausschusssquote = as_teile * 100 / total_parts

                    current_directory = os.path.dirname(os.path.abspath(__file__))
                    process_folder = os.path.join(current_directory, '..', 'Werk', 'Prozesse', process)
                    file_path = os.path.join(process_folder, f'{process}_DS.csv')
                    variant = 'Ausschusssquote'
                    write_kpi(file_path, variant, ausschusssquote)

                    return ausschusssquote
    except:
        return 0


def calculate_fehlproduktionsquote(process, batch, variant, process_df, fehler_excel_path):  # Prozentzahl
    today = process_df.tail(1)["date"].iloc[0]
    processes_today = process_df[process_df["date"] == today]
    failed_processes = process_df[(process_df["date"] == today) & (process_df["exit_code"] == 2)]

    num_failed = failed_processes.shape[0]
    num_total = processes_today.shape[0]

    fehlerproduktionsquote = num_failed / num_total

    current_directory = os.path.dirname(os.path.abspath(__file__))
    process_folder = os.path.join(current_directory, '..', 'Werk', 'Prozesse', process)
    file_path = os.path.join(process_folder, f'{process}_DS.csv')
    variant = 'Fehlproduktionsquote'
    write_kpi(file_path, variant, int(100*fehlerproduktionsquote))

    return fehlerproduktionsquote


def calculate_qualitaetsgrad(process, batch, variant, process_df, fehler_excel_path):  # Prozentzahl
    qualitaetsgrad = 1 - calculate_fehlproduktionsquote(process, batch, variant, process_df, fehler_excel_path)
    current_directory = os.path.dirname(os.path.abspath(__file__))
    process_folder = os.path.join(current_directory, '..', 'Werk', 'Prozesse', process)
    file_path = os.path.join(process_folder, f'{process}_DS.csv')
    variant = 'Qualitaetsgrad'
    write_kpi(file_path, variant, qualitaetsgrad)
    write_kpi(file_path, "OEE - Quality", qualitaetsgrad)

    return qualitaetsgrad


def calculate_yield(process, variant, process_df):
    today = process_df.tail(1)["date"].iloc[0]

    # Filter the DataFrame for the desired variant and today's date
    filtered_df = process_df[(process_df['variant'] == variant) & (process_df['date'] == today) &
                             (process_df['exit_code'] == 0)]

    yeld = filtered_df["menge"].mean()

    current_directory = os.path.dirname(os.path.abspath(__file__))
    process_folder = os.path.join(current_directory, '..', 'Werk', 'Prozesse', process)
    file_path = os.path.join(process_folder, f'{process}_DS.csv')
    variant = 'Leistung/Ausbringung/Yield'
    write_kpi(file_path, variant, yeld)

    return yeld


def calculate_work_in_process(process, variant, process_df, timestamp):
    """
    today = process_df.tail(1)["date"].iloc[0]
    # Events in the last 24 hours with exit code 1, that is, those that are still running
    filtered_process_df = process_df[(process_df["date"] == today) & (process_df["variant"] == variant) & (process_df["exit_code"] == 1)]

    # Get the 'box_capacity' value
    batch = process_df["menge"].tail(1).iloc[0]
    # Return number of pieces

    wip = filtered_process_df["menge"].sum()*batch
    """

    # Redo this if possible
    current_directory = os.path.dirname(os.path.abspath(__file__))
    process_folder = os.path.join(current_directory, '..', 'Werk', 'Prozesse', process)
    file_path = os.path.join(process_folder, f'{process}_DS.csv')
    variant = 'Work in Process'
    write_kpi(file_path, variant, 0)

    return 0


def calculate_oee_av(process, fehler_excel_path):
    downtime = calculate_production_downtime(process, fehler_excel_path)
    # Use datetime.strptime to parse the time string
    time_obj = datetime.strptime(downtime, "%H:%M:%S")
    # Extract the hours, minutes, and seconds and calculate the total seconds
    downtime_seconds = time_obj.hour * 3600 + time_obj.minute * 60 + time_obj.second
    operating_time = 28800  # 8h per day
    availability = (operating_time - downtime_seconds) / operating_time if operating_time > 0 else 0

    current_directory = os.path.dirname(os.path.abspath(__file__))
    process_folder = os.path.join(current_directory, '..', 'Werk', 'Prozesse', process)
    file_path = os.path.join(process_folder, f'{process}_DS.csv')
    variant = 'OEE - Availability/Verfuegbarkeit'
    write_kpi(file_path, variant, availability)

    return availability


def calculate_oee_pe(process, variant, batch, process_df):
    today = process_df.tail(1)["date"].iloc[0]

    # Filter the DataFrame for the desired variant and today's date
    filtered_df = process_df[(process_df['variant'] == variant) & (process_df['date'] == today)]

    # Count the number of rows in the filtered DataFrame and calculate the total runtime
    number_of_runs = filtered_df.shape[0]
    list_of_runs = filtered_df["duration"].tolist()
    runtime = "00:00:00"
    for i in range(number_of_runs):
        runtime = add_timestamps(runtime, list_of_runs[i])
    # Divide by the batch to get the cycle time
    runtime = divide_time(runtime, batch)

    # Use datetime.strptime to parse the time string
    time_obj = datetime.strptime(runtime, "%H:%M:%S")
    # Extract the hours, minutes, and seconds and calculate the total seconds
    runtime_seconds = time_obj.hour * 3600 + time_obj.minute * 60 + time_obj.second

    # Obtain the ideal cycle time in seconds and compare
    if "FK" in variant:
        variant = "FK"
    process_variant = process + variant
    ideal_cycle_times = {"SaegenFS1": 27, "SaegenFS2": 41, "DrehenFK": 97.5, "FraesenFB1": 130.5, "FraesenFB2": 151,
                         "WaschenFB1": 28, "WaschenFB2": 28, "WaschenFK": 28, "MessenFB1": 9, "MessenFB2": 23,
                         "MessenFK": 43, "MontageFB1": 38.5, "MontageFB2": 44}

    ideal_runtime = ideal_cycle_times[process_variant] * number_of_runs

    try:
        performance = min(1, ideal_runtime / runtime_seconds)
    except ZeroDivisionError:
        performance = 1

    current_directory = os.path.dirname(os.path.abspath(__file__))
    process_folder = os.path.join(current_directory, '..', 'Werk', 'Prozesse', process)
    file_path = os.path.join(process_folder, f'{process}_DS.csv')
    variant = 'OEE - Performance/Leistung'
    write_kpi(file_path, variant, performance)

    return performance


"""
def calculate_oee_qa(process, variant, batch, process_df, total_parts_produced, fehler_excel_path):  # Make a DF with the ideal cycle time per variant in each process
    # Calculate Quality
    good_parts_produced = total_parts_produced - calculate_fehlproduktionsquote(process, batch, variant, process_df,  fehler_excel_path)

    quality = good_parts_produced / total_parts_produced if total_parts_produced > 0 else 0

    current_directory = os.path.dirname(os.path.abspath(__file__))
    process_folder = os.path.join(current_directory, '..', 'Werk', 'Prozesse', process)
    file_path = os.path.join(process_folder, f'{process}_DS.csv')
    variant = 'OEE - Quality'
    write_kpi(file_path, variant, quality)

    return quality
"""


def calculate_oeestern(process, total_parts, variant, process_df, timestamp):
    working_time = 8 * 3600

    try:
        oeestern = total_parts * calculate_average_cycle_time(variant, process, process_df, timestamp) / working_time

        current_directory = os.path.dirname(os.path.abspath(__file__))
        process_folder = os.path.join(current_directory, '..', 'Werk', 'Prozesse', process)
        file_path = os.path.join(process_folder, f'{process}_DS.csv')
        variant = 'OEE*'
        write_kpi(file_path, variant, oeestern)

        return oeestern
    except:
        current_directory = os.path.dirname(os.path.abspath(__file__))
        process_folder = os.path.join(current_directory, '..', 'Werk', 'Prozesse', process)
        file_path = os.path.join(process_folder, f'{process}_DS.csv')
        variant = 'OEE*'
        write_kpi(file_path, variant, 0.73)
        return 0.73


def calculate_productivity(process, fehler_excel_path):
    try:
        # Percentage of uptime I think
        productivity = 1 - calculate_production_downtime(process, fehler_excel_path) - \
                       calculate_unscheduled_downtime(process, fehler_excel_path)

        current_directory = os.path.dirname(os.path.abspath(__file__))
        # construct the path to the CSV file
        process_folder = os.path.join(current_directory, '..', 'Werk', 'Prozesse', process)
        file_path = os.path.join(process_folder, f'{process}_DS.csv')
        variant = 'Produktivitaet'
        write_kpi(file_path, variant, productivity)

        return productivity
    except:
        return 0.8


def calculate_losgroesse(process, process_df, variant):
    batch = process_df["menge"].tail(1).iloc[0]

    current_directory = os.path.dirname(os.path.abspath(__file__))

    # Construct the path to the CSV file
    process_folder = os.path.join(current_directory, '..', 'Werk', 'Prozesse', process)
    file_path = os.path.join(process_folder, f'{process}_DS.csv')

    # Formatted variant
    translate = {"FK1": "KS_1", "FK2": "KS_2", "FK3": "KS_3", "FK4": "KS_4", "FK5": "KS_5", "FK6": "KS_6",
                 "FK7": "KS_7", "FK8": "KS_8", "FK": "KS (Ø)", "FB1": "D25", "FB2": "D40", "FS1": "D25", "FS2": "D40"}
    variant = 'Losgroesse ' + f'{translate[variant]}'
    write_kpi(file_path, variant, batch)

    return batch


def calculate_anzahl_typ(process, process_df, variant):
    adjustment = {"Fraesen": 2, "Saegen": 2, "Drehen": 8, "Waschen": 10, "Messen": 10, "Montage": 2}

    current_directory = os.path.dirname(os.path.abspath(__file__))
    # construct the path to the CSV file
    process_folder = os.path.join(current_directory, '..', 'Werk', 'Prozesse', process)
    file_path = os.path.join(process_folder, f'{process}_DS.csv')
    variant = 'Anzahl Typ'
    write_kpi(file_path, variant, adjustment[process])

    return adjustment[process]


"""
DATENSPEICHERUNG-FUNKTIONEN
"""


def write_kpi(file_path, variant, KPI_value):
    # Load the Excel file
    # workbook = openpyxl.load_workbook(file_path)
    ds_df = pd.read_csv(file_path)
    print(ds_df)
    print(variant)
    print("\n\n\n\n\n")
    column_name = variant
    row_index = 0
    new_value = KPI_value
    ds_df.at[row_index, column_name] = new_value
    ds_df.to_csv(file_path, index=False)


def update_global_kpis():
    werk_ds_path = os.path.join("", "..\\Werk\\Werk_DS.csv")
    werk_ds = pd.read_csv(werk_ds_path)

    # Process times and VAR can be calculated directly from the values in Werk_DS
    # Wartezeiten must be converted to seconds (86400 seconds per day)
    wait1 = int(86400 * float(werk_ds.iloc[9]["SF Besprechung"]))
    process1 = int(werk_ds.iloc[2]["Pausen"])
    wait2 = int(86400 * (float(werk_ds.iloc[10]["SF Besprechung"]) + float(werk_ds.iloc[11]["SF Besprechung"])))
    process2 = int(werk_ds.iloc[3]["Pausen"]) + int(werk_ds.iloc[4]["Pausen"]) + int(werk_ds.iloc[5]["Pausen"])
    wait3 = int(86400 * float(werk_ds.iloc[12]["SF Besprechung"]))
    process3 = int(werk_ds.iloc[6]["Pausen"])
    wait4 = int(86400 * (float(werk_ds.iloc[13]["SF Besprechung"]) + float(werk_ds.iloc[14]["SF Besprechung"]) +
                         float(werk_ds.iloc[15]["SF Besprechung"]) + float(werk_ds.iloc[16]["SF Besprechung"]) +
                         float(werk_ds.iloc[17]["SF Besprechung"])))
    process4 = int(werk_ds.iloc[7]["Pausen"])

    # Calculate value-added and non-value-added time in seconds
    va_time = process1 + process2 + process3 + process4
    nva_time = wait1 + wait2 + wait3 + wait4
    var = int((100 * va_time) / (va_time + nva_time))  # In percentage

    # Ausbringung and lead time can also be calculated directly from Werk_DS
    lead_time = va_time + nva_time
    ausbringung_list = [int(float(werk_ds.iloc[i]["Stueckzahl"])) for i in range(2, 8)]
    ausbringung = sum(ausbringung_list)
    print(ausbringung)
    try:
        produktionstakt = int(100 * lead_time / ausbringung_list[5]) / 100
    except ZeroDivisionError:
        produktionstakt = 0

    # Calculating downtime and Fehlproduktionsquote requires opening the process_DS files
    process_list = ["Saegen", "Fraesen", "Drehen", "Waschen", "Messen", "Montage"]
    downtime_list = []
    fehlproduktionsquote_list = []
    productivity = 0  # To be replaced by Performance from Montage

    for process in process_list:
        process_path = os.path.join("", f"..\\Werk\\Prozesse\\{process}\\{process}_DS.csv")
        process_ds = pd.read_csv(process_path)

        # Calculate downtime in seconds
        downtime = str(process_ds.iloc[0]["Production Downtime"])
        try:
            time_obj = datetime.strptime(downtime, "%H:%M:%S")
            downtime = time_obj.hour * 3600 + time_obj.minute * 60 + time_obj.second
            downtime_list.append(downtime)
        except:
            downtime_list.append(0)

        # Calculate Fehlproduktionsquote
        fehlproduktionsquote = process_ds.iloc[0]["Fehlproduktionsquote"]
        fehlproduktionsquote_list.append(fehlproduktionsquote)

        # Obtain productivity
        if process == "Montage":
            productivity = int(100*float(process_ds.iloc[0]["OEE - Performance/Leistung"]))

    global_downtime = sum(downtime_list)
    total_failed_parts = sum([fehlproduktionsquote_list[i] * ausbringung_list[i] for i in range(6)])
    global_fehlproduktionsquote = int(100 * total_failed_parts / ausbringung)  # In percentage

    # Write all those values to the top row of Werk_DS
    werk_ds.at[0, "Ausfallzeit"] = global_downtime
    werk_ds.at[0, "Ausbringung(montage)"] = ausbringung
    werk_ds.at[0, "Produktivitaet"] = productivity
    werk_ds.at[0, "Fehlproduktionsquote"] = global_fehlproduktionsquote
    werk_ds.at[0, "VA Time"] = va_time
    werk_ds.at[0, "NVA Time"] = nva_time
    werk_ds.at[0, "Durchlaufzeit"] = lead_time
    werk_ds.at[0, "VA Ratio"] = var
    werk_ds.at[0, "Wait1"] = wait1
    werk_ds.at[0, "Process1"] = process1
    werk_ds.at[0, "Wait2"] = wait2
    werk_ds.at[0, "Process2"] = process2
    werk_ds.at[0, "Wait3"] = wait3
    werk_ds.at[0, "Process3"] = process3
    werk_ds.at[0, "Wait4"] = wait4
    werk_ds.at[0, "Process4"] = process4
    werk_ds.at[0, "Produktionstakt"] = produktionstakt

    # Save the adjustments based on the standards
    werk_ds.at[0, "Schichtlaenge"] = 60
    werk_ds.at[0, "Pausen"] = 6
    werk_ds.at[0, "SF Besprechung"] = 6
    werk_ds.at[0, "Stueckzahl"] = 48
    werk_ds.at[0, "Kundentakt"] = 60

    # Save the Werk_DS
    werk_ds.to_csv(werk_ds_path, index=False)


def calculate_process_kpis(process, variant, date, timestamp):
    # Excel file paths
    fehler_excel_path = 'Digital_Button.xlsm'
    FabrikVerbindung = pd.read_excel('FabrikVerbindung.xlsx')

    # Obtain the process.csv as a pandas DF
    current_directory = os.path.dirname(os.path.abspath(__file__))
    process_path = os.path.join(current_directory, f"..\\Werk\\Prozesse\\{process}\\{process}" + ".csv")
    process_df = pd.read_csv(process_path)

    # Call every function
    losgroesse = calculate_losgroesse(process, process_df, variant)
    batch = losgroesse
    fehlproduktionsquote = calculate_fehlproduktionsquote(process, batch, variant, process_df, fehler_excel_path)
    qualitaetsgrad = calculate_qualitaetsgrad(process, batch, variant, process_df, fehler_excel_path)
    ausschussquote = calculate_ausschussquote(process, variant, process_df, fehler_excel_path)
    nacharbeitsquote = calculate_nacharbeitquote(process, variant, process_df, fehler_excel_path)
    average_cycle_time = calculate_average_cycle_time(variant, process, process_df, timestamp)
    average_process_time = calculate_process_time(process, variant, process_df, timestamp)
    production_downtime = calculate_production_downtime(process, fehler_excel_path)
    unscheduled_downtime = calculate_unscheduled_downtime(process, fehler_excel_path)
    leistung = calculate_yield(process, variant, process_df)
    work_in_process = calculate_work_in_process(process, variant, process_df, timestamp)
    productivity = calculate_productivity(process, fehler_excel_path)
    anzahl_typ = calculate_anzahl_typ(process, process_df, variant)

    # OEE CALCULATION
    oee_av = calculate_oee_av(process, fehler_excel_path)
    oee_pe = calculate_oee_pe(process, variant, batch, process_df)
    # oee_qa = calculate_oee_qa(process, variant, batch, process_df, total_parts_produced, fehler_excel_path)
    oee = oee_av * oee_pe * qualitaetsgrad
    # oeestern = calculate_oeestern(process, total_parts_produced, variant, process_df, timestamp)

    current_directory = os.path.dirname(os.path.abspath(__file__))
    process_folder = os.path.join(current_directory, '..', 'Werk', 'Prozesse', process)
    file_path = os.path.join(process_folder, f'{process}_DS.csv')
    write_kpi(file_path, "OEE", oee)

    # ##############################################################################
    # Append the calculated values as a row for the process_HistLog.csv (on top)
    hist_log_path = os.path.join("", f"..\\Werk\\Prozesse\\{process}\\" + f"{process}_HistLog.csv")
    ds_path = os.path.join("", f"..\\Werk\\Prozesse\\{process}\\" + f"{process}_DS.csv")

    histlog_row = {"Date": [date], "Time": [timestamp], "Variant": [variant], "Menge": [batch], "OEE Gesamt": [oee],
                   "OEE Verfuegbarkeit": [oee_av], "OEE Leistung": [oee_pe], "OEE Qualitaet": [qualitaetsgrad]}
    histlog_to_append = pd.DataFrame(histlog_row)

    histlog_df = pd.read_csv(hist_log_path)
    histlog_df = pd.concat([histlog_to_append, histlog_df]).reset_index(drop=True)
    histlog_df.to_csv(hist_log_path, index=False)

    # ########################################################################
    # Update Werk_DS
    werk_ds_path = os.path.join("", "..\\Werk\\Werk_DS.csv")
    werk_ds = pd.read_csv(werk_ds_path)
    process_list = ["Saegen", "Fraesen", "Drehen", "Waschen", "Messen", "Montage"]
    process_index = 2 + process_list.index(process)

    """
    # Reformat cycle/process times as seconds
    time_obj = datetime.strptime(average_cycle_time, "%H:%M:%S")
    average_cycle_time = time_obj.hour * 3600 + time_obj.minute * 60 + time_obj.second
    time_obj = datetime.strptime(average_process_time, "%H:%M:%S")
    average_process_time = time_obj.hour * 3600 + time_obj.minute * 60 + time_obj.second
    """

    # This uses the column names at the top of Werk_DS
    werk_ds.at[process_index, "Schichtlaenge"] = average_cycle_time
    werk_ds.at[process_index, "Pausen"] = average_process_time
    werk_ds.at[process_index, "SF Besprechung"] = int(oee * 100)
    werk_ds.at[process_index, "Stueckzahl"] = leistung
    werk_ds.at[process_index, "Kundentakt"] = anzahl_typ
    werk_ds.at[process_index, "Ausfallzeit"] = work_in_process
    werk_ds.to_csv(werk_ds_path, index=False)

    # Recalculate front page KPIs
    update_global_kpis()
