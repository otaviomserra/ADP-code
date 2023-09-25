import os
from openpyxl import load_workbook
import pandas as pd


class ExcelWriter:
    def __init__(self, process_name):
        self.process_name = process_name
        self.excel_file_path = self.get_excel_file_path()

    def get_excel_file_path(self):
        # Construct the path to the Excel file based on your description
        current_directory = os.path.dirname(__file__)  # Assuming this code is in the same directory as the Excel file
        excel_file_path = os.path.join(current_directory, '..', 'Werk', 'Prozesse', 'KPIS_Prozessen_Matrix.xlsx')
        return excel_file_path

    def write_to_cell(self, value):
        try:
            workbook = load_workbook(self.excel_file_path)
            worksheet = workbook.active

            # Assuming "Work in Process" is in column A (adjust as needed)
            row_index = None
            for row in worksheet.iter_rows(min_col=1, max_col=1, values_only=True):
                if row[0] == "Work in Process":
                    row_index = row[0]
                    break

            if row_index is not None:
                # Find the column with the same name as self.process_name (case-sensitive)
                column_index = None
                for cell in worksheet.iter_cols(min_row=1, max_row=1, values_only=True):
                    if cell[0] == self.process_name:
                        column_index = cell[0]
                        break

                if column_index is not None:
                    # Write the specified value to the cell
                    worksheet[column_index + row_index] = value
                    workbook.save(self.excel_file_path)
                # No need to handle cases where the column or row isn't found, as you requested to ignore errors.
        except Exception:
            pass  # Ignore all exceptions
